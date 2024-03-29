import datetime
from urllib import response
from uuid import RFC_4122
from django.shortcuts import get_object_or_404, render, redirect
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from chatbotapp.functions.menuFormatting import makeWeekendReply, menuFormat
from chatbotapp.functions.is_vacation import is_holiday
from chatbotapp.functions.weekdayConverter import weekConverter
from .kakaojsonformat.response import (
    insert_image,
    insert_text,
    make_reply,
    insert_replies,
)
from .form import *
import json
from datetime import date, timedelta
from openpyxl import load_workbook

# Create your views here.


@csrf_exempt
def menu_list(request):
    try:
        menus = RDB.objects.filter(date=date.today())[0]
        return render(
            request, template_name="menu_list.html", context={"menus": menus}
        )
    except:
        menus = RDB.objects.all()
        return render(
            request, template_name="menu_list.html", context={"menus": menus}
        )


@csrf_exempt
def chaSeDae_create(request):
    if request.method == "POST":
        form = chaSeDaeForm(request.POST, request.FILES)
        if form.is_valid():
            menu = form.save()
            return redirect("chatbot:list")
    else:
        form = chaSeDaeForm()
    ctx = {"form": form}
    return render(request, template_name="menu_form.html", context=ctx)


@csrf_exempt
def nano_create(request):
    if request.method == "POST":
        form = nanoForm(request.POST, request.FILES)
        if form.is_valid():
            menu = form.save()
            return redirect("chatbot:list")
    else:
        form = nanoForm()
    ctx = {"form": form}
    return render(request, template_name="menu_form.html", context=ctx)


@csrf_exempt
def RDB_create(request):
    if request.method == "POST":
        form = RDBForm(request.POST, request.FILES)
        if form.is_valid():
            menu = form.save()
            return redirect("chatbot:list")
    else:
        form = RDBForm()
    ctx = {"form": form}
    return render(request, template_name="menu_form.html", context=ctx)


@csrf_exempt
def get_chaSeDae(request):
    answer = request.body.decode("utf-8")
    return_json_str = json.loads(answer)
    return_str = return_json_str["userRequest"]["utterance"]
    # 오늘이 일요일인데 금요일을 호출한다
    # 30일인데 28일꺼 호출
    #  오늘 기준으로
    # 호출한 요일 - 오늘 요일 값을 오늘날짜에 날짜 연산으로 더해주면됌
    # delta 는 오늘기준으로 사용자가 선택한요일의 차이값입니다

    # selectedDay 는 사용자가 선택한 요일에 상응하는 날짜입니다

    if return_str == "차세대융합기술원" or return_str == "🚗🚗🚗차세대융합기술원":

        if is_holiday():
            response = insert_text("공휴일에는 식단을 제공하지 않습니다😊\n행복한 하루 되세요")
            response = makeWeekendReply("차세대융합기술원", response)
            return JsonResponse(response)

        text = "품절 확인 기능이 추가되었습니다\n실시간으로 업데이트됩니다\n식당 방문전 확인 부탁드립니다🤩\n오늘 차세대융합기술원 식단\n\n"
        menu = ChaSeDae.objects.filter(date=date.today())[0]

        text += menuFormat("[맘스]", menu.moms)
        text += menuFormat("[셰프]", menu.chef)
        text += menuFormat("[정찬]", menu.special)
        text += menuFormat("[샐러드]", menu.salad)
        text += menuFormat("[저녁]", menu.dinner)
        text += menuFormat("[TakeOut]", menu.takeOut)

        # 실제 보여줄 음식에 대한 메뉴는 위에서 처리했다 이 밑에는 이제 사용자의 클릭을 유도하는 메뉴 생성
        response = insert_text(text)
        response = makeWeekendReply("차세대융합기술원", response)

        return JsonResponse(response)

    elif return_str == "CSO" or return_str == "cso":
        text = "원하시는 기능의 버튼을 클릭해주세요!"

        menu = ChaSeDae.objects.filter(date=date.today())[0]
        print(menu.moms.find("품절"))
        response = insert_text(text)
        if menu.moms.find("품절") == -1:
            reply = make_reply("맘스품절", "차세대융합기술원맘스품절")
            response = insert_replies(response, reply)
        else:
            reply = make_reply("맘스품절해제", "차세대융합기술원맘스품절해제")
            response = insert_replies(response, reply)
        if menu.chef.find("품절") == -1:
            reply = make_reply("셰프품절", "차세대융합기술원셰프품절")
            response = insert_replies(response, reply)
        else:
            reply = make_reply("셰프품절해제", "차세대융합기술원셰프품절해제")
            response = insert_replies(response, reply)
        if menu.special.find("품절") == -1:
            reply = make_reply("정찬품절", "차세대융합기술원정찬품절")
            response = insert_replies(response, reply)
        else:
            reply = make_reply("정찬품절해제", "차세대융합기술원정찬품절")
            response = insert_replies(response, reply)

        return JsonResponse(response)

    elif return_str == "차세대융합기술원맘스품절":
        text = "차세대융합기술원 맘스 메뉴가 품절되었습니다"
        menu = ChaSeDae.objects.filter(date=date.today())[0]
        menu.moms = "❗️품절❗️," + menu.moms
        menu.save()
        response = insert_text(text)
        return JsonResponse(response)

    elif return_str == "차세대융합기술원셰프품절":
        text = "차세대융합기술원 셰프 메뉴가 품절되었습니다"
        menu = ChaSeDae.objects.filter(date=date.today())[0]
        menu.chef = "❗️품절❗️," + menu.chef
        menu.save()
        response = insert_text(text)
        return JsonResponse(response)

    elif return_str == "차세대융합기술원정찬품절":
        text = "차세대융합기술원 정찬 메뉴가 품절되었습니다"
        menu = ChaSeDae.objects.filter(date=date.today())[0]
        menu.special = "❗️품절❗️," + menu.special
        menu.save()
        response = insert_text(text)
        return JsonResponse(response)

    elif return_str == "차세대융합기술원맘스품절해제":
        text = "차세대융합기술원 맘스 메뉴 품절이 해제되었습니다"
        menu = ChaSeDae.objects.filter(date=date.today())[0]
        menu.moms = menu.moms.replace("❗️품절❗️,", "")
        menu.save()
        response = insert_text(text)
        return JsonResponse(response)

    elif return_str == "차세대융합기술원셰프품절해제":
        text = "차세대융합기술원 셰프 메뉴 품절이 해제되었습니다"
        menu = ChaSeDae.objects.filter(date=date.today())[0]
        menu.chef = menu.chef.replace("❗️품절❗️,", "")
        menu.save()
        response = insert_text(text)
        return JsonResponse(response)

    elif return_str == "차세대융합기술원정찬품절해제":
        text = "차세대융합기술원 셰프 메뉴 품절이 해제되었습니다"
        menu = ChaSeDae.objects.filter(date=date.today())[0]
        menu.special = menu.special.replace("❗️품절❗️,", "")
        menu.save()
        response = insert_text(text)
        return JsonResponse(response)

    elif return_str == "차세대융합기술원문의사항":
        text = (
            "📪 문의사항 : 우혜림 영양사 [prefla@naver.com] \n" "031-888-9497 로 연락 바랍니다."
        )
        response = insert_text(text)
        response = makeWeekendReply("차세대융합기술원", response)
        return JsonResponse(response)

    elif return_str == "차세대융합기술원식단표":
        delta = weekConverter("월") - datetime.datetime.today().weekday()
        selectedDay = date.today() + timedelta(days=delta)
        table = MenuTables.objects.filter(date=selectedDay, restaurant="차세대")[
            0
        ]
        response = insert_image(
            request.get_host() + "/media/" + str(table.table_img), return_str
        )
        response = makeWeekendReply("차세대융합기술원", response)
        return JsonResponse(response)

    else:
        delta = weekConverter(return_str) - datetime.datetime.today().weekday()
        selectedDay = date.today() + timedelta(days=delta)
        menu = ChaSeDae.objects.filter(date=selectedDay)[0]
        text = selectedDay.strftime("%m-%d")
        text += " 차세대융합기술원식단\n\n"

        text += menuFormat("[맘스]", menu.moms)
        text += menuFormat("[셰프]", menu.chef)
        text += menuFormat("[정찬]", menu.special)
        text += menuFormat("[샐러드]", menu.salad)
        text += menuFormat("[저녁]", menu.dinner)
        text += menuFormat("[TakeOut]", menu.takeOut)

        response = insert_text(text)
        response = makeWeekendReply("차세대융합기술원", response)
        return JsonResponse(response)


@csrf_exempt
def get_nano(request):
    answer = request.body.decode("utf-8")
    return_json_str = json.loads(answer)
    return_str = return_json_str["userRequest"]["utterance"]

    if return_str == "한국나노기술원" or return_str == "🍚한국나노기술원":
        # 여기에 데이터 베이스에서 차세대 융합기술원에서 하루 전체 메뉴가져오는 로직 지금은 text 로 dummy 로 쓰겠음
        if is_holiday():
            response = insert_text("공휴일에는 식단을 제공하지 않습니다😊\n행복한 하루 되세요")
            response = makeWeekendReply("한국나노기술원", response)
            return JsonResponse(response)

        try:
            menu = Nano.objects.filter(date=date.today())[0]
            text = "품절 확인 기능이 추가되었습니다\n실시간으로 업데이트됩니다\n식당 방문전 확인 부탁드립니다🤩\n오늘 한국나노기술원 식단\n\n"

            text += menuFormat("[정성이 가득한 점심 A코너]", menu.lunchA)
            text += menuFormat("[정성이 가득한 점심 B코너]", menu.lunchB)
            text += menuFormat("[PLUS]", menu.plus)
            text += menuFormat("[하루를 마무리 하는 저녁]", menu.dinner)

            # 실제 보여줄 음식에 대한 메뉴는 위에서 처리했다 이 밑에는 이제 사용자의 클릭을 유도하는 메뉴 생성
            response = insert_text(text)
            response = makeWeekendReply("한국나노기술원", response)

            return JsonResponse(response)
        except:
            response = insert_text("아직 식단이 제공되지 않았어요!")
            response = makeWeekendReply("한국나노기술원", response)
            return JsonResponse(response)

    elif return_str == "NSO" or return_str == "nso":
        text = "원하시는 기능의 버튼을 클릭해주세요!"

        menu = Nano.objects.filter(date=date.today())[0]
        response = insert_text(text)
        if menu.lunchA.find("품절") == -1:
            reply = make_reply("A코너품절", "한국나노기술원A코너품절")
            response = insert_replies(response, reply)
        else:
            reply = make_reply("A코너품절해제", "한국나노기술원A코너품절해제")
            response = insert_replies(response, reply)
        if menu.lunchB.find("품절") == -1:
            reply = make_reply("B코너품절", "한국나노기술원B코너품절")
            response = insert_replies(response, reply)
        else:
            reply = make_reply("B코너품절해제", "한국나노기술원B코너품절해제")
            response = insert_replies(response, reply)

        return JsonResponse(response)

    elif return_str == "한국나노기술원A코너품절":
        text = "한국나노기술원 A코스가 품절되었습니다"
        menu = Nano.objects.filter(date=date.today())[0]
        menu.lunchA = "❗️품절❗️," + menu.lunchA
        menu.save()
        response = insert_text(text)
        return JsonResponse(response)

    elif return_str == "한국나노기술원B코너품절":
        text = "한국나노기술원 B코스가 품절되었습니다"
        menu = Nano.objects.filter(date=date.today())[0]
        menu.lunchB = "❗️품절❗️," + menu.lunchB
        menu.save()
        response = insert_text(text)
        return JsonResponse(response)

    elif return_str == "한국나노기술원A코너품절해제":
        text = "한국나노기술원 A코너 품절이 해제되었습니다"
        menu = Nano.objects.filter(date=date.today())[0]
        menu.lunchA = menu.lunchA.replace("❗️품절❗️,", "")
        menu.save()
        response = insert_text(text)
        return JsonResponse(response)

    elif return_str == "한국나노기술원B코너품절해제":
        text = "한국나노기술원 B코너 품절이 해제되었습니다"
        menu = Nano.objects.filter(date=date.today())[0]
        menu.lunchB = menu.lunchB.replace("❗️품절❗️,", "")
        menu.save()
        response = insert_text(text)
        return JsonResponse(response)

    elif return_str == "한국나노기술원문의사항":
        text = (
            "⏰ 운영시간안내\n- 중식 11:30 ~ 13:10\n"
            "- 석식 17:30 ~ 18:30\n\n"
            "📃 상기 식단은 시장 동향에 따라 변경될 수 있는 점 양해 바랍니다.\n\n"
            "📜 원산지 표시\n"
            "쌀(국내산), 배추김치(맛김치 - 배추: 국내산, 고춧가루: 중국산)"
            "깍두기(무 : 국내산, 고춧가루 : 중국산) 우육(호주산,미국산), 돈육"
            "(국내산,수입산), 계육(국산) 식육가공품(국내산/수입산)\n\n"
            "제공되는 메뉴 및 원산지는 식자재 수급 현황에 따라 변경 될 수 있으니"
            "정확한 정보는 식당 입구에 게시된 일일메뉴표를 참고바랍니다."
        )
        response = insert_text(text)
        response = makeWeekendReply("한국나노기술원", response)
        return JsonResponse(response)

    elif return_str == "한국나노기술원식단표":
        delta = weekConverter("월") - datetime.datetime.today().weekday()
        selectedDay = date.today() + timedelta(days=delta)
        table = MenuTables.objects.filter(date=selectedDay, restaurant="나노")[0]
        response = insert_image(
            request.get_host() + "/media/" + str(table.table_img), return_str
        )
        response = makeWeekendReply("한국나노기술원", response)
        return JsonResponse(response)

    else:
        delta = weekConverter(return_str) - datetime.datetime.today().weekday()
        selectedDay = date.today() + timedelta(days=delta)
        try:
            menu = Nano.objects.filter(date=selectedDay)[0]

            text = selectedDay.strftime("%m-%d")
            text += " 한국나노기술원식단\n\n"

            text += menuFormat("[정성이 가득한 점심 A코너]", menu.lunchA)
            text += menuFormat("[정성이 가득한 점심 B코너]", menu.lunchB)
            text += menuFormat("[PLUS]", menu.plus)
            text += menuFormat("[하루를 마무리 하는 저녁]", menu.dinner)

            response = insert_text(text)
            response = makeWeekendReply("한국나노기술원", response)

            return JsonResponse(response)
        except:
            text = selectedDay.strftime("%m-%d")
            text += " 아직 식단이 제공되지 않았어요!\n"
            response = insert_text(text)
            response = makeWeekendReply("한국나노기술원", response)
            return JsonResponse(response)


@csrf_exempt
def get_R_DB(request):
    answer = request.body.decode("utf-8")
    return_json_str = json.loads(answer)
    return_str = return_json_str["userRequest"]["utterance"]

    if return_str == "경기 RDB" or return_str == "🍙경기 RDB":

        if is_holiday():
            response = insert_text("공휴일에는 식단을 제공하지 않습니다😊\n행복한 하루 되세요")
            response = makeWeekendReply("경기 RDB", response)
            return JsonResponse(response)
        text = "품절 확인 기능이 추가되었습니다\n실시간으로 업데이트됩니다\n식당 방문전 확인 부탁드립니다🤩\n오늘 경기RDB 식단\n\n"
        try:
            menu = RDB.objects.filter(date=date.today())[0]
        except:
            response = insert_text("이번주 식단이 아직 도착하지 않았어요, 최대한 빨리 업데이트 하겠습니다.")
            response = makeWeekendReply("경기 RDB", response)
            return JsonResponse(response)

        text += menuFormat("[한식]", menu.korea)
        text += menuFormat("[일품]", menu.special)
        text += menuFormat("[점심 플러스바]", menu.lunch_plus)
        text += menuFormat("[석식]", menu.dinner)
        text += menuFormat("[저녁 플러스바]", menu.dinner_plus)
        text += menuFormat("[TaktOut]", menu.takeOut)

        # 실제 보여줄 음식에 대한 메뉴는 위에서 처리했다 이 밑에는 이제 사용자의 클릭을 유도하는 메뉴 생성
        response = insert_text(text)
        response = makeWeekendReply("경기 RDB", response)

        return JsonResponse(response)

    elif return_str == "RSO" or return_str == "rso":
        text = "원하시는 기능의 버튼을 클릭해주세요!"

        menu = RDB.objects.filter(date=date.today())[0]
        response = insert_text(text)
        if menu.korea.find("품절") == -1:
            reply = make_reply("한식품절", "경기 RDB한식품절")
            response = insert_replies(response, reply)
        else:
            reply = make_reply("한식품절해제", "경기 RDB한식품절해제")
            response = insert_replies(response, reply)
        if menu.special.find("품절") == -1:
            reply = make_reply("일품품절", "경기 RDB일품품절")
            response = insert_replies(response, reply)
        else:
            reply = make_reply("일품품절해제", "경기 RDB일품품절해제")
            response = insert_replies(response, reply)

        return JsonResponse(response)

    elif return_str == "경기 RDB한식품절":
        text = "경기 RDB 한식메뉴가 품절되었습니다"
        menu = RDB.objects.filter(date=date.today())[0]
        menu.korea = "❗️품절❗️," + menu.korea
        menu.save()
        response = insert_text(text)
        return JsonResponse(response)

    elif return_str == "경기 RDB일품품절":
        text = "경기 RDB 일품메뉴가 품절되었습니다"
        menu = RDB.objects.filter(date=date.today())[0]
        menu.special = "❗️품절❗️," + menu.special
        menu.save()
        response = insert_text(text)
        return JsonResponse(response)

    elif return_str == "경기 RDB한식품절해제":
        text = "경기 RDB 한식메뉴 품절이 해제되었습니다"
        menu = RDB.objects.filter(date=date.today())[0]
        menu.korea = menu.korea.replace("❗️품절❗️,", "")
        menu.save()
        response = insert_text(text)
        return JsonResponse(response)

    elif return_str == "경기 RDB일품품절해제":
        text = "경기 RDB 일품메뉴 품절이 해제되었습니다"
        menu = RDB.objects.filter(date=date.today())[0]
        menu.special = menu.special.replace("❗️품절❗️,", "")
        menu.save()
        response = insert_text(text)
        return JsonResponse(response)

    elif return_str == "경기 RDB문의사항":
        text = (
            "📪 문의사항 : 조혜성 영양사 [hyeseong92@daum.net] \n"
            "010-3168-9547 로 연락 바랍니다."
        )
        response = insert_text(text)
        response = makeWeekendReply("경기 RDB", response)
        return JsonResponse(response)

    elif return_str == "경기 RDB식단표":
        delta = weekConverter("월") - datetime.datetime.today().weekday()
        selectedDay = date.today() + timedelta(days=delta)
        table = MenuTables.objects.filter(date=selectedDay, restaurant="RDB")[
            0
        ]
        response = insert_image(
            request.get_host() + "/media/" + str(table.table_img), return_str
        )
        response = makeWeekendReply("경기 RDB", response)
        return JsonResponse(response)

    else:
        delta = weekConverter(return_str) - datetime.datetime.today().weekday()
        selectedDay = date.today() + timedelta(days=delta)
        menu = RDB.objects.filter(date=selectedDay)[0]
        text = selectedDay.strftime("%m-%d")
        text += " 경기 RDB 식단\n\n"

        text += menuFormat("[한식]", menu.korea)
        text += menuFormat("[일품]", menu.special)
        text += menuFormat("[점심 플러스바]", menu.lunch_plus)
        text += menuFormat("[석식]", menu.dinner)
        text += menuFormat("[저녁 플러스바]", menu.dinner_plus)
        text += menuFormat("[TaktOut]", menu.takeOut)

        response = insert_text(text)
        response = makeWeekendReply("경기 RDB", response)
        return JsonResponse(response)


@csrf_exempt
def get_etc(request):
    answer = request.body.decode("utf-8")
    return_json_str = json.loads(answer)
    return_str = return_json_str["userRequest"]["utterance"]

    if return_str == "📬건의사항":
        openurl = "https://open.kakao.com/o/si30eZVd"
        response = insert_text(
            "⁉️오류제보 / 기능 건의 ⁉️\n {}\n 링크를 클릭후 \n 편하게 채팅해주세요\n 여러분들의 오류제보가 \n 융밥을 더 성장시킵니다".format(
                openurl
            )
        )
        # reply = make_reply("🏡홈으로", "홈")
        # answer = insert_replies(answer, reply)
        return JsonResponse(response)


@csrf_exempt
def add_rdb(request):
    file = request.FILES["myfile"]
    print(file)
    # data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
    load_wb = load_workbook(file, data_only=True)
    # 시트 이름으로 불러오기
    load_ws = load_wb["경기알앤디비"]

    # 셀 주소로 값 출력
    a = 65
    a = "D4"
    for i in range(68, 73):

        tempKorea = (
            load_ws[chr(i) + "4"].value
            + ","
            + load_ws[chr(i) + "5"].value
            + ","
            + load_ws[chr(i) + "6"].value
            + ","
            + load_ws[chr(i) + "7"].value
            + ","
            + load_ws[chr(i) + "8"].value
            + ","
            + load_ws[chr(i) + "9"].value
        )
        tempSpecial = (
            load_ws[chr(i) + "10"].value
            + ","
            + load_ws[chr(i) + "11"].value
            + ","
            + load_ws[chr(i) + "12"].value
            + ","
            + load_ws[chr(i) + "13"].value
            + ","
            + load_ws[chr(i) + "14"].value
            + ","
            + load_ws[chr(i) + "15"].value
        )
        tempLunchPlus = (
            load_ws[chr(i) + "16"].value
            + ","
            + load_ws[chr(i) + "17"].value
            + ","
            + load_ws[chr(i) + "18"].value
            + ","
            + load_ws[chr(i) + "19"].value
        )
        tempDinner = (
            load_ws[chr(i) + "20"].value
            + ","
            + load_ws[chr(i) + "21"].value
            + ","
            + load_ws[chr(i) + "22"].value
            + ","
            + load_ws[chr(i) + "23"].value
            + load_ws[chr(i) + "24"].value
        )
        tempDinnerPlus = (
            load_ws[chr(i) + "25"].value
            + ","
            + load_ws[chr(i) + "26"].value
            + ","
            + load_ws[chr(i) + "27"].value
        )
        tempTakeOut = (
            load_ws[chr(i) + "28"].value + "," + load_ws[chr(i) + "29"].value
        )
        RDB.objects.create(
            date=load_ws[chr(i) + "3"].value,
            korea=tempKorea,
            special=tempSpecial,
            lunch_plus=tempLunchPlus,
            dinner=tempDinner,
            dinner_plus=tempDinnerPlus,
            takeOut=tempTakeOut,
        )
    text = ""
    response = insert_text(text)
    return JsonResponse(response)


def xstr(s):
    return "" if s is None else str(s)


def uploadFile(request):
    if request.method == "POST":
        # Fetching the form data
        fileTitle = request.POST["fileTitle"]
        uploadedFile = request.FILES["uploadedFile"]
        if fileTitle == "rdb":
            load_wb = load_workbook(uploadedFile, data_only=True)
            load_ws = load_wb["경기알앤디비"]
            for i in range(68, 73):
                if (
                    load_ws[chr(i) + "4"].value == None
                    or load_ws[chr(i) + "5"].value == None
                ):
                    continue
                tempKorea = (
                    load_ws[chr(i) + "4"].value
                    + ","
                    + load_ws[chr(i) + "5"].value
                    + ","
                    + load_ws[chr(i) + "6"].value
                    + ","
                    + load_ws[chr(i) + "7"].value
                    + ","
                    + load_ws[chr(i) + "8"].value
                    + ","
                    + load_ws[chr(i) + "9"].value
                )
                tempSpecial = (
                    load_ws[chr(i) + "10"].value
                    + ","
                    + load_ws[chr(i) + "11"].value
                    + ","
                    + load_ws[chr(i) + "12"].value
                    + ","
                    + load_ws[chr(i) + "13"].value
                    + ","
                    + load_ws[chr(i) + "14"].value
                    + ","
                    + load_ws[chr(i) + "15"].value
                )
                tempLunchPlus = (
                    load_ws[chr(i) + "16"].value
                    + ","
                    + load_ws[chr(i) + "17"].value
                    + ","
                    + load_ws[chr(i) + "18"].value
                    + ","
                    + load_ws[chr(i) + "19"].value
                )
                tempDinner = (
                    load_ws[chr(i) + "20"].value
                    + ","
                    + load_ws[chr(i) + "21"].value
                    + ","
                    + load_ws[chr(i) + "22"].value
                    + ","
                    + load_ws[chr(i) + "23"].value
                    + load_ws[chr(i) + "24"].value
                )
                tempDinnerPlus = (
                    load_ws[chr(i) + "25"].value
                    + ","
                    + load_ws[chr(i) + "26"].value
                )
                tempTakeOut = (
                    load_ws[chr(i) + "27"].value
                    + ","
                    + load_ws[chr(i) + "28"].value
                )
                RDB.objects.create(
                    date=load_ws[chr(i) + "3"].value,
                    korea=tempKorea,
                    special=tempSpecial,
                    lunch_plus=tempLunchPlus,
                    dinner=tempDinner,
                    dinner_plus=tempDinnerPlus,
                    takeOut=tempTakeOut,
                )
        elif fileTitle == "cha":
            prev = 68
            load_wb = load_workbook(uploadedFile, data_only=True)
            load_ws = load_wb["융기원"]
            for i in range(68, 73):
                if (
                    load_ws[chr(i) + "4"].value == None
                    or load_ws[chr(i) + "5"].value == None
                ):
                    continue
                tempMoms = (
                    load_ws[chr(i) + "4"].value
                    + ","
                    + load_ws[chr(i) + "5"].value
                    + ","
                    + load_ws[chr(i) + "6"].value
                    + ","
                    + load_ws[chr(i) + "7"].value
                    + ","
                    + load_ws[chr(i) + "8"].value
                    + ","
                    + load_ws[chr(i) + "9"].value
                )
                tempChef = (
                    load_ws[chr(i) + "10"].value
                    + ","
                    + load_ws[chr(i) + "11"].value
                    + ","
                    + load_ws[chr(i) + "12"].value
                    + ","
                    + load_ws[chr(i) + "13"].value
                    + ","
                    + load_ws[chr(i) + "14"].value
                    + ","
                    + load_ws[chr(i) + "15"].value
                )
                tempSalad = (
                    load_ws[chr(i) + "16"].value
                    + ","
                    + load_ws[chr(i) + "17"].value
                    + ","
                    + load_ws[chr(i) + "18"].value
                )
                tempSpecial = (
                    load_ws[chr(i) + "19"].value
                    + ","
                    + load_ws[chr(i) + "20"].value
                    + ","
                    + load_ws[chr(i) + "21"].value
                    + ","
                    + load_ws[chr(i) + "22"].value
                    + ","
                    + load_ws[chr(i) + "23"].value
                    + ","
                    + load_ws[chr(i) + "24"].value
                    + ","
                    + load_ws[chr(i) + "25"].value
                    + ","
                    + load_ws[chr(i) + "26"].value
                )
                tempDinner = (
                    load_ws[chr(i) + "27"].value
                    + ","
                    + load_ws[chr(i) + "28"].value
                    + ","
                    + load_ws[chr(i) + "29"].value
                    + ","
                    + load_ws[chr(i) + "30"].value
                    + load_ws[chr(i) + "31"].value
                    + load_ws[chr(i) + "32"].value
                )
                if load_ws[chr(i) + "35"].value != None:
                    prev = i
                try:
                    tempTakeOut = (
                        load_ws[chr(i) + "33"].value
                        + ","
                        + load_ws[chr(i) + "34"].value
                        + ","
                        + load_ws[chr(i) + "35"].value
                    )
                except:
                    tempTakeOut = (
                        load_ws[chr(i) + "33"].value
                        + ","
                        + load_ws[chr(i) + "34"].value
                        + ","
                        + load_ws[chr(prev) + "35"].value
                    )

                ChaSeDae.objects.create(
                    date=load_ws[chr(i) + "3"].value,
                    moms=tempMoms,
                    chef=tempChef,
                    special=tempSpecial,
                    salad=tempSalad,
                    dinner=tempDinner,
                    takeOut=tempTakeOut,
                )

        elif fileTitle == "nano":
            load_wb = load_workbook(uploadedFile, data_only=True)
            load_ws = load_wb["주간식단표(나노)"]
            # 메뉴의 숫자가 딱 안맞을 경우에 None 타입이 나오게 되고 None 타입과 String 붙히려면 오류나서 각 메뉴의 멘 마지막 행은 예외처리해줌
            for i in range(66, 71):
                if (
                    load_ws[chr(i) + "5"].value == None
                    or load_ws[chr(i) + "6"].value == None
                ):
                    continue
                tempLunchA = (
                    load_ws[chr(i) + "5"].value
                    + ","
                    + load_ws[chr(i) + "6"].value
                    + ","
                    + load_ws[chr(i) + "7"].value
                    + ","
                    + load_ws[chr(i) + "8"].value
                    + ","
                    + load_ws[chr(i) + "9"].value
                    + ","
                    + xstr(load_ws[chr(i) + "10"].value)
                )
                tempLunchB = (
                    load_ws[chr(i) + "11"].value
                    + ","
                    + load_ws[chr(i) + "12"].value
                    + ","
                    + load_ws[chr(i) + "13"].value
                    + ","
                    + load_ws[chr(i) + "14"].value
                    + ","
                    + xstr(load_ws[chr(i) + "15"].value)
                )
                tempPlus = load_ws[chr(i) + "16"].value

                tempDinner = (
                    load_ws[chr(i) + "17"].value
                    + ","
                    + load_ws[chr(i) + "18"].value
                    + ","
                    + load_ws[chr(i) + "19"].value
                    + ","
                    + load_ws[chr(i) + "20"].value
                    + ","
                    + load_ws[chr(i) + "21"].value
                    + ","
                    + xstr(load_ws[chr(i) + "22"].value)
                )

                Nano.objects.create(
                    date=load_ws[chr(i) + "4"].value,
                    lunchA=tempLunchA,
                    lunchB=tempLunchB,
                    plus=tempPlus,
                    dinner=tempDinner,
                )

    documents = Document.objects.all()

    return render(request, "upload_file.html", context={"files": documents})
